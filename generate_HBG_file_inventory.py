import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Resolve the project root whether this file is stored in the project root or scripts.
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = (
    SCRIPT_DIR.parent
    if SCRIPT_DIR.name.casefold() == "scripts"
    else SCRIPT_DIR
)

ROOT_PATH = PROJECT_ROOT / "output" / "processed_hbg"
REPORTS_PATH = PROJECT_ROOT / "output" / "reports" / "generate_HBG_file_inventory"
OUTPUT_FILE_PREFIX = "file_inventory"

INVENTORY_COLUMNS = [
    "folder_name",
    "relative_subfolder",
    "file_name",
    "file_extension",
    "file_path",
    "file_size_MB",
    "guessed_type",
    "cad_view",
    "cad_selected",
    "use_for_gemini",
    "note",
]

IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg"}
RAW_CAD_EXTENSIONS = {".iam", ".ipt", ".stp", ".step"}
EXCEL_EXTENSIONS = {".xlsx", ".xls", ".csv"}

COMPACT_CAD_VIEWS = {
    "TLF": frozenset({"top", "left", "front"}),
    "TRF": frozenset({"top", "right", "front"}),
    "TLB": frozenset({"top", "left", "back"}),
    "TRB": frozenset({"top", "right", "back"}),
    "BLF": frozenset({"bottom", "left", "front"}),
    "BRF": frozenset({"bottom", "right", "front"}),
    "BLB": frozenset({"bottom", "left", "back"}),
    "BRB": frozenset({"bottom", "right", "back"}),
}

DIRECTION_WORDS = {
    "top": {"top", "oben"},
    "bottom": {"bottom", "unten"},
    "left": {"left", "links", "linke", "linker"},
    "right": {"right", "rechts", "rechte", "rechter"},
    "front": {"front", "vorne", "vorn", "vorder"},
    "back": {"back", "rear", "hinten", "rueck", "ruck", "rück"},
}

DIRECTION_ORDER = ("top", "bottom", "left", "right", "front", "back")


def contains_any(text: str, keywords: tuple[str, ...]) -> bool:
    return any(keyword in text for keyword in keywords)


def normalize_search_text(text: str) -> str:
    """Normalize common filename separators without changing German characters."""
    return " ".join(text.casefold().replace("_", " ").replace("-", " ").split())


def is_dfc_structure_screenshot(relative_path_text: str, extension: str) -> bool:
    if extension not in IMAGE_EXTENSIONS:
        return False

    normalized_text = normalize_search_text(relative_path_text)
    has_structure_word = contains_any(
        normalized_text,
        ("struktur", "strucktur", "structure", "strukturbaum", "baugruppenstruktur"),
    )
    has_dfc_reference = "dfc" in normalized_text
    has_strong_structure_name = contains_any(
        normalized_text, ("strukturbaum", "baugruppenstruktur")
    )

    return (has_dfc_reference and has_structure_word) or has_strong_structure_name


def guess_file_type(file_path: Path, baugruppe_folder: Path) -> str:
    """Guess the file role from its name, parent folders, and extension."""
    extension = file_path.suffix.lower()
    relative_path_text = str(file_path.relative_to(baugruppe_folder)).casefold()

    if contains_any(relative_path_text, ("cad-zugriff", "cad_zugriff", "cad zugriff")):
        return "CAD_access_ignore"

    # DFC structure screenshots must be checked before generic image detection.
    if is_dfc_structure_screenshot(relative_path_text, extension):
        return "DFC_structure_screenshot"

    if extension == ".idw":
        return "IDW_drawing"

    # Specific document names take priority over generic PDF/Excel extensions.
    if contains_any(
        relative_path_text,
        ("stückliste", "stueckliste", "bom", "teileliste", "part list", "parts list"),
    ):
        return "BOM"

    if contains_any(relative_path_text, ("datenblatt", "datasheet", "data sheet")):
        return "Datenblatt"

    if contains_any(
        relative_path_text,
        ("zeichnung", "drawing", "technical drawing", "technische zeichnung"),
    ):
        return "Technische_Zeichnung"

    if extension in RAW_CAD_EXTENSIONS:
        return "CAD_raw"

    if extension in IMAGE_EXTENSIONS or contains_any(
        relative_path_text, ("screenshot", "view", "ansicht", "cad")
    ):
        return "CAD_screenshot"

    if extension in EXCEL_EXTENSIONS:
        return "Excel_unknown"

    if extension == ".pdf":
        return "Technische_Zeichnung"

    return "unknown"


def default_gemini_usage(guessed_type: str) -> str:
    if guessed_type == "Technische_Zeichnung":
        return "priority_1_candidate"
    if guessed_type in {
        "DFC_structure_screenshot",
        "BOM",
        "Datenblatt",
        "Excel_unknown",
    }:
        return "priority_2_candidate"
    return "nein"


def detect_cad_view(file_name: str) -> tuple[str, frozenset[str]]:
    """Return a normalized CAD view name and its viewing directions."""
    stem = Path(file_name).stem
    compact_match = re.search(
        r"(?:^|[^A-Z0-9])(TLF|TRF|TLB|TRB|BLF|BRF|BLB|BRB)(?:$|[^A-Z0-9])",
        stem.upper(),
    )
    if compact_match:
        directions = COMPACT_CAD_VIEWS[compact_match.group(1)]
    else:
        words = set(re.findall(r"[a-zA-ZäöüÄÖÜß]+", stem.casefold()))
        directions = frozenset(
            direction
            for direction, aliases in DIRECTION_WORDS.items()
            if words.intersection(aliases)
        )

    view_name = "_".join(
        direction for direction in DIRECTION_ORDER if direction in directions
    )
    return view_name, directions


def select_cad_screenshots(dataframe: pd.DataFrame) -> pd.DataFrame:
    """Select up to three useful CAD screenshot views per Baugruppe."""
    dataframe["cad_view"] = ""
    dataframe["cad_selected"] = ""

    cad_rows = dataframe[dataframe["guessed_type"] == "CAD_screenshot"]
    for _, group in cad_rows.groupby("folder_name", sort=False):
        candidates = []
        for index, row in group.iterrows():
            view_name, directions = detect_cad_view(str(row["file_name"]))
            dataframe.at[index, "cad_view"] = view_name
            candidates.append(
                {
                    "index": index,
                    "directions": directions,
                    "size": float(row["file_size_MB"]),
                    "name": str(row["file_name"]).casefold(),
                }
            )

        selected = []

        def add_best(predicate, preference=None) -> None:
            eligible = [
                candidate
                for candidate in candidates
                if candidate not in selected and predicate(candidate["directions"])
            ]
            if not eligible or len(selected) >= 3:
                return

            def selection_key(candidate):
                preference_score = (
                    preference(candidate["directions"]) if preference else 0
                )
                return (
                    -preference_score,
                    -len(candidate["directions"]),
                    -candidate["size"],
                    candidate["name"],
                )

            selected.append(sorted(eligible, key=selection_key)[0])

        # Start with an informative upper-front diagonal view.
        add_best(
            lambda directions: {
                "top",
                "front",
            }.issubset(directions)
            and bool({"left", "right"}.intersection(directions))
        )

        # Then choose an opposite lower-rear diagonal view where available.
        first_directions = selected[0]["directions"] if selected else frozenset()

        def opposite_side_score(directions: frozenset[str]) -> int:
            if "left" in first_directions and "right" in directions:
                return 1
            if "right" in first_directions and "left" in directions:
                return 1
            return 0

        add_best(
            lambda directions: {
                "bottom",
                "back",
            }.issubset(directions)
            and bool({"left", "right"}.intersection(directions)),
            opposite_side_score,
        )

        # A straight front view usually exposes the assembly layout clearly.
        add_best(lambda directions: directions == frozenset({"front"}))

        # Fill missing slots with the most informative and diverse remaining views.
        while len(selected) < min(3, len(candidates)):
            remaining = [candidate for candidate in candidates if candidate not in selected]
            covered_directions = (
                set().union(*(candidate["directions"] for candidate in selected))
                if selected
                else set()
            )
            remaining.sort(
                key=lambda candidate: (
                    -len(candidate["directions"] - covered_directions),
                    -len(candidate["directions"]),
                    -candidate["size"],
                    candidate["name"],
                )
            )
            selected.append(remaining[0])

        selected_indices = {candidate["index"] for candidate in selected}
        for candidate in candidates:
            index = candidate["index"]
            is_selected = index in selected_indices
            dataframe.at[index, "cad_selected"] = "ja" if is_selected else "nein"
            dataframe.at[index, "use_for_gemini"] = (
                "priority_1_candidate" if is_selected else "nein"
            )

    return dataframe


def scan_dataset(root_path: Path) -> tuple[pd.DataFrame, list[Path]]:
    root_path = root_path.expanduser().resolve()
    if not root_path.exists():
        raise FileNotFoundError(f"Dataset folder does not exist: {root_path}")
    if not root_path.is_dir():
        raise NotADirectoryError(f"Dataset path is not a folder: {root_path}")

    # Every first-level directory is treated as one Baugruppe.
    baugruppe_folders = sorted(
        (path for path in root_path.iterdir() if path.is_dir()),
        key=lambda path: path.name.casefold(),
    )

    rows = []
    for baugruppe_folder in baugruppe_folders:
        files = sorted(
            (path for path in baugruppe_folder.rglob("*") if path.is_file()),
            key=lambda path: str(path).casefold(),
        )

        for file_path in files:
            # Ignore temporary Microsoft Office lock files.
            if file_path.name.startswith("~$"):
                continue

            try:
                size_mb = round(file_path.stat().st_size / (1024 * 1024), 3)
            except OSError as error:
                print(f"Warning: could not read file metadata: {file_path} ({error})")
                continue

            relative_parent = file_path.parent.relative_to(baugruppe_folder)
            relative_subfolder = "" if relative_parent == Path(".") else str(relative_parent)
            guessed_type = guess_file_type(file_path, baugruppe_folder)

            rows.append(
                {
                    "folder_name": baugruppe_folder.name,
                    "relative_subfolder": relative_subfolder,
                    "file_name": file_path.name,
                    "file_extension": file_path.suffix.lower(),
                    "file_path": str(file_path.resolve()),
                    "file_size_MB": size_mb,
                    "guessed_type": guessed_type,
                    "cad_view": "",
                    "cad_selected": "",
                    "use_for_gemini": default_gemini_usage(guessed_type),
                    "note": "",
                }
            )

    dataframe = pd.DataFrame(rows, columns=INVENTORY_COLUMNS)
    dataframe = select_cad_screenshots(dataframe)
    return dataframe, baugruppe_folders


def save_inventory(dataframe: pd.DataFrame, output_path: Path) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        dataframe.to_excel(writer, sheet_name="file_inventory", index=False)
        worksheet = writer.sheets["file_inventory"]

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        preferred_widths = {
            "folder_name": 22,
            "relative_subfolder": 28,
            "file_name": 42,
            "file_extension": 15,
            "file_path": 70,
            "file_size_MB": 14,
            "guessed_type": 26,
            "cad_view": 24,
            "cad_selected": 14,
            "use_for_gemini": 20,
            "note": 35,
        }
        for column_number, column_name in enumerate(INVENTORY_COLUMNS, start=1):
            column_letter = get_column_letter(column_number)
            worksheet.column_dimensions[column_letter].width = preferred_widths[column_name]
            worksheet.cell(row=1, column=column_number).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        worksheet.column_dimensions["F"].width = 14
        for cell in worksheet["F"][1:]:
            cell.number_format = "0.000"


def print_statistics(dataframe: pd.DataFrame, baugruppe_folders: list[Path]) -> None:
    print(f"Baugruppe folders scanned: {len(baugruppe_folders)}")
    print(f"Files recorded: {len(dataframe)}")
    print("\nFiles by guessed_type:")

    if dataframe.empty:
        print("  No files found.")
        return

    counts = dataframe["guessed_type"].value_counts().sort_index()
    for guessed_type, count in counts.items():
        print(f"  {guessed_type}: {count}")


def build_inventory(root_path: Path) -> Path:
    root_path = root_path.expanduser().resolve()
    dataframe, baugruppe_folders = scan_dataset(root_path)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    reports_path = REPORTS_PATH.expanduser().resolve()
    reports_path.mkdir(parents=True, exist_ok=True)
    output_path = reports_path / f"{OUTPUT_FILE_PREFIX}_{timestamp}.xlsx"

    save_inventory(dataframe, output_path)
    print_statistics(dataframe, baugruppe_folders)
    print(f"\nInventory created: {output_path}")
    return output_path


if __name__ == "__main__":
    build_inventory(ROOT_PATH)
