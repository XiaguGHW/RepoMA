from pathlib import Path
import datetime
import logging
import re
import shutil

import pandas as pd
from openpyxl import load_workbook

try:
    from tqdm import tqdm
except ImportError:
    def tqdm(iterable, **kwargs):
        return iterable

from Google_native_connector_multiplefiles_1 import BoschLLMConnector


# === KONFIGURATION ===
# ================================================================================

SCRIPT_ORDNER = Path(__file__).resolve().parent
INPUT_ORDNER = SCRIPT_ORDNER / "input"
OUTPUT_ORDNER = SCRIPT_ORDNER / "output"

ASSEMBLY_DATEI = INPUT_ORDNER / "all HBG random no label.xlsx"
FUNCTIONAL_CLASSES_DATEI = INPUT_ORDNER / "Functional_classes.xlsx"

# Diesen Pfad auf dem Firmenrechner anpassen.
# In diesem Ordner liegen die einzelnen Baugruppenordner.
BAUGRUPPEN_DATEN_PFAD = Path(r"C:\HIER\PFAD\ZU\BAUGRUPPEN_DATEN")

SAP_SPALTE = "SAP-Nummer"
TC_SPALTE = "Teamcenter ID"
BENENNUNG_EN_SPALTE = "Benennung (EN)"
BENENNUNG_DE_SPALTE = "Benennung (DE)"
FUNKTIONSKLASSE_SPALTE = "Funktionsklasse"
ERGEBNIS_SPALTE = "Gemini_Label"

# Fuer den ersten Test werden nur die ersten 10 Baugruppen verarbeitet.
# Fuer alle Baugruppen: ANZAHL_ZU_TESTEN = None
ANZAHL_ZU_TESTEN = 10

BOSCH_FARM_API_KEY = "HIER_API_KEY_EINTRAGEN"
MODEL_NAME = "gemini-2.5-pro"

GENERATION_CONFIG = {
    "temperature": 0.1,
    "topP": 0.95,
    "candidateCount": 1,
    "maxOutputTokens": 1000,
    "stopSequences": []
}

# Nur diese Dateitypen werden direkt an Gemini uebergeben.
DIREKT_UNTERSTUETZTE_ENDUNGEN = {
    ".pdf",
    ".txt", ".md", ".csv", ".json", ".xml",
    ".jpg", ".jpeg", ".png", ".webp"
}

# Excel-Dateien werden nicht konvertiert und nicht an Gemini uebergeben.
EXCEL_ENDUNGEN = {".xlsx", ".xls", ".xlsm", ".xlsb"}

MAX_DATEIGROESSE_MB = 20


# --- Logging-Konfiguration ---
OUTPUT_ORDNER.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(
            OUTPUT_ORDNER / "verarbeitung_baugruppen_dateien.log",
            mode="w",
            encoding="utf-8"
        ),
        logging.StreamHandler()
    ]
)


def zellwert_als_text(wert):
    if wert is None or pd.isna(wert):
        return ""

    text = str(wert).strip()

    if text.lower() == "nan":
        return ""

    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]

    return text


def normalisiere_fuer_suche(wert):
    text = zellwert_als_text(wert).upper()
    return re.sub(r"[^A-Z0-9]", "", text)


def lade_daten():
    if not ASSEMBLY_DATEI.exists():
        raise FileNotFoundError(f"Assembly-Datei nicht gefunden: {ASSEMBLY_DATEI}")

    if not FUNCTIONAL_CLASSES_DATEI.exists():
        raise FileNotFoundError(
            f"Functional-Classes-Datei nicht gefunden: {FUNCTIONAL_CLASSES_DATEI}"
        )

    assembly_df = pd.read_excel(
        ASSEMBLY_DATEI,
        dtype=str,
        keep_default_na=False
    )
    classes_df = pd.read_excel(
        FUNCTIONAL_CLASSES_DATEI,
        dtype=str,
        keep_default_na=False
    )

    benoetigte_spalten = {
        SAP_SPALTE,
        TC_SPALTE,
        BENENNUNG_EN_SPALTE,
        BENENNUNG_DE_SPALTE
    }
    fehlende_spalten = benoetigte_spalten.difference(assembly_df.columns)

    if fehlende_spalten:
        raise ValueError(
            "Fehlende Spalten in der Assembly-Datei: "
            + ", ".join(sorted(fehlende_spalten))
        )

    if FUNKTIONSKLASSE_SPALTE not in classes_df.columns:
        raise ValueError(
            f"Spalte '{FUNKTIONSKLASSE_SPALTE}' fehlt in "
            f"{FUNCTIONAL_CLASSES_DATEI.name}."
        )

    klassen_liste = []
    bereits_gesehen = set()

    for wert in classes_df[FUNKTIONSKLASSE_SPALTE]:
        klasse = zellwert_als_text(wert)
        if klasse and klasse not in bereits_gesehen:
            klassen_liste.append(klasse)
            bereits_gesehen.add(klasse)

    if not klassen_liste:
        raise ValueError("Keine Funktionsklassen gefunden.")

    logging.info("Assembly-Datei geladen: %s Zeilen", len(assembly_df))
    logging.info("Funktionsklassen geladen: %s", len(klassen_liste))

    return assembly_df, klassen_liste


def erstelle_ordner_index():
    if not BAUGRUPPEN_DATEN_PFAD.exists():
        raise FileNotFoundError(
            f"Baugruppen-Datenpfad nicht gefunden: {BAUGRUPPEN_DATEN_PFAD}"
        )

    ordner_index = []

    for ordner in BAUGRUPPEN_DATEN_PFAD.rglob("*"):
        if ordner.is_dir():
            normalisierter_name = normalisiere_fuer_suche(ordner.name)
            if normalisierter_name:
                ordner_index.append((ordner, normalisierter_name))

    logging.info("Baugruppenordner indexiert: %s", len(ordner_index))
    return ordner_index


def finde_baugruppenordner(sap_nummer, tc_nummer, ordner_index):
    suchnummern = {
        nummer
        for nummer in (
            normalisiere_fuer_suche(sap_nummer),
            normalisiere_fuer_suche(tc_nummer)
        )
        if len(nummer) >= 4
    }

    if not suchnummern:
        return None, "Keine gueltige SAP- oder Teamcenter-Nummer"

    treffer = []

    for ordner, normalisierter_name in ordner_index:
        if any(nummer in normalisierter_name for nummer in suchnummern):
            treffer.append(ordner)

    eindeutige_treffer = sorted(set(treffer), key=lambda pfad: str(pfad).lower())

    if len(eindeutige_treffer) == 1:
        return eindeutige_treffer[0], "OK"

    if not eindeutige_treffer:
        return None, "Kein passender Ordner gefunden"

    logging.warning(
        "Mehrere Ordner gefunden fuer SAP=%s / TC=%s: %s",
        zellwert_als_text(sap_nummer),
        zellwert_als_text(tc_nummer),
        "; ".join(str(pfad) for pfad in eindeutige_treffer)
    )
    return None, f"Mehrere passende Ordner gefunden ({len(eindeutige_treffer)})"


def finde_unterstuetzte_dateien(baugruppenordner):
    unterstuetzte_dateien = []
    uebersprungene_dateien = []

    for datei in sorted(
        baugruppenordner.rglob("*"),
        key=lambda pfad: str(pfad).lower()
    ):
        if not datei.is_file():
            continue

        relativer_pfad = datei.relative_to(baugruppenordner)
        endung = datei.suffix.lower()

        if datei.name.startswith("~$"):
            uebersprungene_dateien.append((relativer_pfad, "Temporaere Datei"))
            continue

        if any(teil.startswith(".") for teil in relativer_pfad.parts):
            uebersprungene_dateien.append((relativer_pfad, "Versteckte Datei"))
            continue

        if endung in EXCEL_ENDUNGEN:
            uebersprungene_dateien.append((relativer_pfad, "Excel uebersprungen"))
            continue

        if endung not in DIREKT_UNTERSTUETZTE_ENDUNGEN:
            uebersprungene_dateien.append((relativer_pfad, "Dateityp nicht unterstuetzt"))
            continue

        dateigroesse = datei.stat().st_size

        if dateigroesse == 0:
            uebersprungene_dateien.append((relativer_pfad, "Leere Datei"))
            continue

        if dateigroesse > MAX_DATEIGROESSE_MB * 1024 * 1024:
            uebersprungene_dateien.append((relativer_pfad, "Datei zu gross"))
            continue

        unterstuetzte_dateien.append(datei.resolve())

    return unterstuetzte_dateien, uebersprungene_dateien


def baue_prompt(zeile, klassen_liste, baugruppenordner, dateien):
    klassen_text = "\n".join(f"- {klasse}" for klasse in klassen_liste)

    if dateien:
        dateien_text = "\n".join(
            f"- {datei.relative_to(baugruppenordner.resolve())}"
            for datei in dateien
        )
    else:
        dateien_text = "- Keine direkt lesbaren Dateien vorhanden"

    prompt = f"""
Du bist ein technischer Experte fuer Baugruppenklassifikation im Maschinen- und Anlagenbau.

Klassifiziere die folgende Baugruppe in genau eine der vorgegebenen Funktionsklassen.
Nutze dafuer die Bezeichnungen und den Inhalt aller hochgeladenen Unterlagen.

SAP-Nummer:
{zellwert_als_text(zeile[SAP_SPALTE])}

Teamcenter ID:
{zellwert_als_text(zeile[TC_SPALTE])}

Benennung (EN):
{zellwert_als_text(zeile[BENENNUNG_EN_SPALTE])}

Benennung (DE):
{zellwert_als_text(zeile[BENENNUNG_DE_SPALTE])}

Hochgeladene Unterlagen:
{dateien_text}

Zulaessige Funktionsklassen:
{klassen_text}

Regeln:
- Waehle genau eine Funktionsklasse aus der Liste.
- Antworte ausschliesslich mit dem exakten Namen der Funktionsklasse.
- Keine Erklaerung, keine zusaetzlichen Woerter.
- Beurteile die Baugruppe insgesamt und nicht nur anhand eines einzelnen Dateinamens.
- Wenn keine Klasse eindeutig passt, antworte mit "Nicht klassifizierbar".

Antwort:
"""
    return prompt


def normalisiere_und_pruefe_antwort(antwort, klassen_liste):
    antwort_text = str(antwort).strip()
    antwort_text = antwort_text.strip("`").strip().strip('"').strip("'").strip()

    gueltige_antworten = list(klassen_liste)
    if "Nicht klassifizierbar" not in gueltige_antworten:
        gueltige_antworten.append("Nicht klassifizierbar")

    if antwort_text in gueltige_antworten:
        return antwort_text

    antworten_nach_kleinschreibung = {
        klasse.casefold(): klasse for klasse in gueltige_antworten
    }
    return antworten_nach_kleinschreibung.get(antwort_text.casefold())


def klassifiziere_baugruppe(llm, zeile, klassen_liste, baugruppenordner, dateien):
    prompt = baue_prompt(
        zeile,
        klassen_liste,
        baugruppenordner,
        dateien
    )

    letzte_antwort = ""

    for versuch in range(1, 3):
        if versuch == 2:
            prompt += (
                "\nDie vorherige Antwort war nicht exakt eine der erlaubten Klassen. "
                "Antworte jetzt nur mit einem exakten Klassennamen aus der Liste."
            )

        antwort = llm.ask_about_files(
            file_paths=[str(datei) for datei in dateien],
            question=prompt,
            generation_config=GENERATION_CONFIG
        )

        letzte_antwort = str(antwort).strip()
        gueltige_antwort = normalisiere_und_pruefe_antwort(
            letzte_antwort,
            klassen_liste
        )

        if gueltige_antwort is not None:
            return gueltige_antwort

        logging.warning(
            "Ungueltige Gemini-Antwort in Versuch %s: %s",
            versuch,
            letzte_antwort
        )

    return f"FEHLER: Ungueltige Gemini-Antwort: {letzte_antwort}"


def bereite_ausgabedatei_vor():
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_pfad = OUTPUT_ORDNER / (
        f"{ASSEMBLY_DATEI.stem}_Gemini_Result_{timestamp}.xlsx"
    )

    shutil.copy2(ASSEMBLY_DATEI, output_pfad)

    workbook = load_workbook(output_pfad)
    worksheet = workbook.active

    vorhandene_spalten = {
        zellwert_als_text(zelle.value): zelle.column
        for zelle in worksheet[1]
        if zellwert_als_text(zelle.value)
    }

    if ERGEBNIS_SPALTE in vorhandene_spalten:
        ergebnis_spaltennummer = vorhandene_spalten[ERGEBNIS_SPALTE]
    else:
        ergebnis_spaltennummer = worksheet.max_column + 1
        worksheet.cell(row=1, column=ergebnis_spaltennummer, value=ERGEBNIS_SPALTE)

    workbook.save(output_pfad)
    return output_pfad, workbook, worksheet, ergebnis_spaltennummer


def main():
    baugruppen_tabelle, funktionsklassen = lade_daten()
    ordner_index = erstelle_ordner_index()

    if ANZAHL_ZU_TESTEN is None:
        zu_verarbeiten = baugruppen_tabelle
    else:
        zu_verarbeiten = baugruppen_tabelle.head(ANZAHL_ZU_TESTEN)

    output_pfad, workbook, worksheet, ergebnis_spalte = bereite_ausgabedatei_vor()

    llm = BoschLLMConnector(
        model_name=MODEL_NAME,
        api_key=BOSCH_FARM_API_KEY
    )

    for index, zeile in tqdm(
        zu_verarbeiten.iterrows(),
        total=len(zu_verarbeiten),
        desc="Baugruppen klassifizieren"
    ):
        sap_nummer = zellwert_als_text(zeile[SAP_SPALTE])
        tc_nummer = zellwert_als_text(zeile[TC_SPALTE])

        baugruppenordner, ordner_status = finde_baugruppenordner(
            sap_nummer,
            tc_nummer,
            ordner_index
        )

        if baugruppenordner is None:
            antwort = f"FEHLER: {ordner_status}"
        else:
            dateien, uebersprungene_dateien = finde_unterstuetzte_dateien(
                baugruppenordner
            )

            excel_anzahl = sum(
                1
                for _, grund in uebersprungene_dateien
                if grund == "Excel uebersprungen"
            )

            logging.info(
                "BG %s / %s | Ordner: %s | Upload: %s | Uebersprungen: %s "
                "(davon Excel: %s)",
                sap_nummer,
                tc_nummer,
                baugruppenordner,
                len(dateien),
                len(uebersprungene_dateien),
                excel_anzahl
            )

            for relativer_pfad, grund in uebersprungene_dateien:
                logging.debug("Uebersprungen: %s | %s", relativer_pfad, grund)

            try:
                antwort = klassifiziere_baugruppe(
                    llm,
                    zeile,
                    funktionsklassen,
                    baugruppenordner,
                    dateien
                )
            except Exception as e:
                logging.exception(
                    "Fehler bei SAP=%s / TC=%s",
                    sap_nummer,
                    tc_nummer
                )
                antwort = f"FEHLER: {e}"

        excel_zeile = int(index) + 2
        worksheet.cell(
            row=excel_zeile,
            column=ergebnis_spalte,
            value=antwort
        )

        # Nach jeder Baugruppe speichern, damit bei einem Abbruch nichts verloren geht.
        workbook.save(output_pfad)

        print(f"{index + 1}/{len(zu_verarbeiten)}")
        print("SAP-Nummer:", sap_nummer)
        print("Teamcenter ID:", tc_nummer)
        print("Gemini-Label:", antwort)
        print("-" * 40)

    workbook.close()
    print("Fertig. Ergebnis gespeichert in:", output_pfad)


if __name__ == "__main__":
    main()
