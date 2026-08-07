"""Create the overview workbook for the workshop classification results.

How to use
----------
1. Set INPUT_DIR below to the absolute path of the folder containing all Excel
   files from the workshop.
2. Run:  python workshop_annotation_auswertung.py

The script reads only the FIRST sheet of every workbook.
It does NOT use the "ID" column, because IDs can occur more than once.
Instead, it matches Baugruppen using "SAP-Nummer" first and, if necessary,
"Teamcenter ID". The output order is taken from one TN_A form (all TN_A
forms use the same order).
"""

from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Settings: only change INPUT_DIR if your folder is stored somewhere else.
# ---------------------------------------------------------------------------
INPUT_DIR = Path(r"C:\Users\YOUR_USERNAME\Documents\Workshop_06.08.2026")
GROUND_TRUTH_FILENAME = "Liste_Workshop_70_Baugruppe_2.Versuch.xlsx"
OUTPUT_FILENAME = "Workshop_Annotation_Auswertung_07.08.2026.xlsx"
PARTICIPANT_PREFIX = "KlassifikationWorkshop_Template_06.08.26_"


def normalise_header(value: Any) -> str:
    """Make header comparisons robust against spaces, hyphens and capitals."""
    text = "" if value is None else str(value)
    return re.sub(r"[^a-z0-9]", "", text.lower())


def normalise_key(value: Any) -> str:
    """Return one stable representation for SAP and Teamcenter identifiers."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def find_column(headers: dict[str, int], *possible_names: str) -> int | None:
    for name in possible_names:
        column = headers.get(normalise_header(name))
        if column is not None:
            return column
    return None


def read_first_sheet(path: Path) -> tuple[list[str], list[list[Any]]]:
    """Read values from the first worksheet only (not any later sheets)."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        raise ValueError(f"The first sheet of '{path.name}' is empty.")

    headers = ["" if value is None else str(value).strip() for value in rows[0]]
    return headers, [list(row) for row in rows[1:] if any(value is not None for value in row)]


def participant_name_from_filename(path: Path) -> str:
    """Use a short, readable participant label as the result-column header."""
    stem = path.stem
    suffix = stem.removeprefix(PARTICIPANT_PREFIX).strip(" _-")
    return suffix or stem


def column_value(row: list[Any], column: int | None) -> Any:
    return row[column] if column is not None and column < len(row) else None


def get_identifier_columns(header_map: dict[str, int], file_name: str) -> tuple[int | None, int | None]:
    """Find the two allowed matching columns in a workshop worksheet."""
    sap_col = find_column(header_map, "SAP-Nummer", "SAP Nummer")
    teamcenter_col = find_column(header_map, "Teamcenter ID")
    if sap_col is None and teamcenter_col is None:
        raise ValueError(
            f"'{file_name}' must contain at least 'SAP-Nummer' or 'Teamcenter ID' "
            "in its first sheet."
        )
    return sap_col, teamcenter_col


def find_matching_row(
    target_row: list[Any],
    target_sap_col: int | None,
    target_teamcenter_col: int | None,
    source_rows: list[list[Any]],
    source_sap_col: int | None,
    source_teamcenter_col: int | None,
    source_name: str,
) -> list[Any] | None:
    """Find one row by SAP number, with Teamcenter ID as fallback/check.

    SAP is the preferred identifier. If it is absent or not unique, Teamcenter
    ID is used. If both identifiers are present, they also disambiguate a
    duplicate SAP or Teamcenter value.
    """
    sap = normalise_key(column_value(target_row, target_sap_col))
    teamcenter = normalise_key(column_value(target_row, target_teamcenter_col))

    sap_matches = [
        row for row in source_rows
        if sap and normalise_key(column_value(row, source_sap_col)) == sap
    ]
    teamcenter_matches = [
        row for row in source_rows
        if teamcenter and normalise_key(column_value(row, source_teamcenter_col)) == teamcenter
    ]

    # With both identifiers, use their intersection when that identifies exactly one row.
    if sap_matches and teamcenter_matches:
        common_matches = [row for row in sap_matches if row in teamcenter_matches]
        if len(common_matches) == 1:
            return common_matches[0]

    if len(sap_matches) == 1:
        return sap_matches[0]
    if len(teamcenter_matches) == 1:
        return teamcenter_matches[0]

    if not sap_matches and not teamcenter_matches:
        return None

    identifiers = f"SAP-Nummer='{sap}' / Teamcenter ID='{teamcenter}'"
    raise ValueError(
        f"More than one matching row was found in '{source_name}' for {identifiers}. "
        "Please check whether SAP-Nummer and Teamcenter ID are filled in consistently."
    )


def main() -> None:
    if not INPUT_DIR.is_dir():
        raise FileNotFoundError(
            "INPUT_DIR does not exist. Please enter the absolute path of your "
            f"workshop folder at the top of this script. Current value: {INPUT_DIR}"
        )

    ground_truth_path = INPUT_DIR / GROUND_TRUTH_FILENAME
    if not ground_truth_path.is_file():
        raise FileNotFoundError(
            f"Ground-truth file not found: '{GROUND_TRUTH_FILENAME}' in {INPUT_DIR}"
        )

    participant_files = sorted(
        path
        for path in INPUT_DIR.glob(f"{PARTICIPANT_PREFIX}*.xlsx")
        if not path.name.startswith("~$")
    )
    if not participant_files:
        raise FileNotFoundError(
            f"No participant files beginning with '{PARTICIPANT_PREFIX}' were found."
        )

    tn_a_files = [path for path in participant_files if "TN_A_" in path.name]
    if not tn_a_files:
        raise ValueError("No TN_A file was found. At least one TN_A form is needed for the output order.")
    order_source_path = tn_a_files[0]

    # Read ground truth. Matching intentionally uses SAP-Nummer / Teamcenter ID,
    # never the ID column.
    gt_headers, gt_rows = read_first_sheet(ground_truth_path)
    gt_header_map = {normalise_header(header): index for index, header in enumerate(gt_headers)}
    gt_sap_col, gt_teamcenter_col = get_identifier_columns(gt_header_map, ground_truth_path.name)
    gt_truth_col = find_column(gt_header_map, "Ground Truth", "GroundTruth")
    if gt_truth_col is None:
        raise ValueError(
            f"'{ground_truth_path.name}' must contain the column 'Ground Truth' in its first sheet."
        )

    # The first TN_A form determines the required display order.
    order_headers, order_rows = read_first_sheet(order_source_path)
    order_header_map = {normalise_header(header): index for index, header in enumerate(order_headers)}
    order_sap_col, order_teamcenter_col = get_identifier_columns(order_header_map, order_source_path.name)
    order_rows = [
        row for row in order_rows
        if normalise_key(column_value(row, order_sap_col))
        or normalise_key(column_value(row, order_teamcenter_col))
    ]
    if len(order_rows) != 70:
        print(f"Warning: TN_A form contains {len(order_rows)} non-empty BG rows (70 were expected).")

    ground_truth_rows: list[list[Any]] = []
    for row_number, order_row in enumerate(order_rows, start=2):
        gt_row = find_matching_row(
            order_row, order_sap_col, order_teamcenter_col,
            gt_rows, gt_sap_col, gt_teamcenter_col, ground_truth_path.name,
        )
        if gt_row is None:
            sap = normalise_key(column_value(order_row, order_sap_col))
            teamcenter = normalise_key(column_value(order_row, order_teamcenter_col))
            raise ValueError(
                f"No Ground Truth row was found for row {row_number} in '{order_source_path.name}' "
                f"(SAP-Nummer='{sap}', Teamcenter ID='{teamcenter}')."
            )
        ground_truth_rows.append(gt_row)

    # Keep every participant's rows. Later, each TN_A row is searched via
    # SAP-Nummer / Teamcenter ID, so TN_B's reversed display order is irrelevant.
    participant_results: list[tuple[str, list[list[Any]], int | None, int | None, int]] = []
    for participant_path in participant_files:
        headers, rows = read_first_sheet(participant_path)
        header_map = {normalise_header(header): index for index, header in enumerate(headers)}
        sap_col, teamcenter_col = get_identifier_columns(header_map, participant_path.name)
        class_col = find_column(header_map, "Baugruppen-Klasse", "Baugruppen Klasse")
        if class_col is None:
            raise ValueError(
                f"'{participant_path.name}' must contain 'Baugruppen-Klasse' in its first sheet."
            )
        participant_results.append(
            (participant_name_from_filename(participant_path), rows, sap_col, teamcenter_col, class_col)
        )

    # These five information columns reproduce the old Übersicht format.
    # If Baugruppenebene is absent in the new ground-truth file, the column is
    # kept empty so that Ground Truth remains in column F.
    info_columns = [
        ("SAP-Nummer", ("SAP-Nummer", "SAP Nummer")),
        ("Teamcenter ID", ("Teamcenter ID",)),
        ("Benennung (EN)", ("Benennung (EN)", "Benennung EN")),
        ("Benennung (DE)", ("Benennung (DE)", "Benennung DE")),
        ("Baugruppenebene", ("Baugruppenebene",)),
    ]
    info_source_columns = [find_column(gt_header_map, *names) for _, names in info_columns]

    output_path = INPUT_DIR / OUTPUT_FILENAME
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Übersicht"

    output_headers = [name for name, _ in info_columns] + ["Ground Truth"]
    output_headers.extend(name for name, *_ in participant_results)
    sheet.append(output_headers)

    for order_row, gt_row in zip(order_rows, ground_truth_rows):
        row_values = [column_value(gt_row, source_col) for source_col in info_source_columns]
        row_values.append(column_value(gt_row, gt_truth_col))
        for participant_name, participant_rows, sap_col, teamcenter_col, class_col in participant_results:
            participant_row = find_matching_row(
                order_row, order_sap_col, order_teamcenter_col,
                participant_rows, sap_col, teamcenter_col, participant_name,
            )
            row_values.append(column_value(participant_row, class_col) if participant_row else None)
        sheet.append(row_values)

    # Simple, readable formatting for the generated overview.
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet.freeze_panes = "G2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 34

    preferred_widths = [18, 20, 32, 32, 18, 24] + [28] * len(participant_results)
    for index, width in enumerate(preferred_widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    workbook.save(output_path)

    print("Finished successfully.")
    print(f"TN_A order source: {order_source_path.name}")
    print(f"Participants included ({len(participant_results)}): " + ", ".join(name for name, *_ in participant_results))
    print(f"Output file: {output_path}")


if __name__ == "__main__":
    try:
        main()
    except Exception as error:
        print(f"ERROR: {error}", file=sys.stderr)
        sys.exit(1)
