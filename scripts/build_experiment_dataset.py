#!/usr/bin/env python3
"""Build the Task 3 classification dataset from Ground Truth and folder coverage.

The script keeps every Ground-Truth row.  A missing folder path stays empty so
that the folder-coverage workbook can be completed later and the dataset can
then be generated again.

Example
-------
python build_experiment_dataset.py \
  --ground-truth-excel 129BG.xlsx \
  --coverage-excel 129BG_folder_coverage_2026-08-31_15-54-18.xlsx
"""

from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


GROUND_TRUTH_SHEET = "Final_Ground_Truth"
FOLDER_MATCHES_SHEET = "folder_matches"
DEFAULT_OUTPUT = "classification_experiment_dataset.xlsx"


def clean_text(value: object) -> str:
    """Return a stable text representation for IDs and labels."""
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalise_header(value: object) -> str:
    return " ".join(clean_text(value).casefold().replace("_", " ").split())


def locate_column(columns: list[object], candidates: tuple[str, ...], *, required: bool) -> object | None:
    """Find a column by an exact normalised name, then by a unique prefix."""
    normalised = {normalise_header(column): column for column in columns}
    for candidate in candidates:
        candidate_normalised = normalise_header(candidate)
        if candidate_normalised in normalised:
            return normalised[candidate_normalised]

    matches = [
        column
        for column in columns
        if any(normalise_header(column).startswith(normalise_header(candidate)) for candidate in candidates)
    ]
    if len(matches) == 1:
        return matches[0]
    if required:
        available = ", ".join(map(str, columns))
        raise ValueError(f"Could not find one of {candidates}. Available columns: {available}")
    return None


def add_matching_key(frame: pd.DataFrame, sap_column: object | None, tc_column: object | None) -> pd.DataFrame:
    """Use SAP first and Teamcenter only if SAP is unavailable."""
    result = frame.copy()
    sap = result[sap_column].map(clean_text) if sap_column is not None else pd.Series("", index=result.index)
    teamcenter = result[tc_column].map(clean_text) if tc_column is not None else pd.Series("", index=result.index)
    result["_match_key"] = sap.where(sap.ne(""), teamcenter)
    return result


def source_register_fills(workbook_path: Path, sheet_name: str, register_header: object) -> dict[str, PatternFill]:
    """Copy the original E1/E2/M cell fills so the new sheet matches it exactly."""
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    worksheet = workbook[sheet_name]
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1))
    register_column = next(
        cell.column
        for cell in header_row
        if normalise_header(cell.value) == normalise_header(register_header)
    )

    fills: dict[str, PatternFill] = {}
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=register_column)
        value = clean_text(cell.value)
        if value in {"E1", "E2", "M"} and value not in fills:
            fills[value] = copy(cell.fill)
    workbook.close()

    fallback = {
        "E1": PatternFill("solid", fgColor="C6EFCE"),
        "E2": PatternFill("solid", fgColor="FFEB9C"),
        "M": PatternFill("solid", fgColor="FFC7CE"),
    }
    return {register: fills.get(register, fallback[register]) for register in fallback}


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--ground-truth-excel", type=Path, required=True)
    parser.add_argument("--coverage-excel", type=Path, required=True)
    parser.add_argument("--ground-truth-sheet", default=GROUND_TRUTH_SHEET)
    parser.add_argument("--folder-matches-sheet", default=FOLDER_MATCHES_SHEET)
    parser.add_argument("--output", type=Path, default=Path(DEFAULT_OUTPUT))
    args = parser.parse_args()

    ground_truth = pd.read_excel(args.ground_truth_excel, sheet_name=args.ground_truth_sheet, dtype=str)
    folder_matches = pd.read_excel(args.coverage_excel, sheet_name=args.folder_matches_sheet, dtype=str)

    gt_sap = locate_column(ground_truth.columns.tolist(), ("SAP-Nummer", "SAP Nummer"), required=False)
    gt_tc = locate_column(ground_truth.columns.tolist(), ("Teamcenter", "Teamcenter ID", "Teamcenter-ID"), required=False)
    if gt_sap is None and gt_tc is None:
        raise ValueError("The Ground Truth sheet needs SAP-Nummer or Teamcenter for matching.")

    gt_english = locate_column(ground_truth.columns.tolist(), ("Benennung (E)",), required=True)
    gt_german = locate_column(ground_truth.columns.tolist(), ("Benennung (D)",), required=True)
    gt_label = locate_column(ground_truth.columns.tolist(), ("Ground Truth",), required=True)
    gt_register = locate_column(ground_truth.columns.tolist(), ("Register",), required=True)
    gt_alternatives = locate_column(
        ground_truth.columns.tolist(), ("Weitere zulässige Ground Truth",), required=False
    )

    match_sap = locate_column(folder_matches.columns.tolist(), ("SAP-Nummer", "SAP Nummer"), required=False)
    match_tc = locate_column(folder_matches.columns.tolist(), ("Teamcenter", "Teamcenter ID", "Teamcenter-ID"), required=False)
    if match_sap is None and match_tc is None:
        raise ValueError("The folder_matches sheet needs SAP-Nummer or Teamcenter for matching.")
    match_path = locate_column(
        folder_matches.columns.tolist(), ("Found_Folder_Paths", "Found Folder Paths"), required=True
    )
    match_file_count = locate_column(
        folder_matches.columns.tolist(), ("Total_File_Count", "Total File Count"), required=False
    )

    ground_truth = add_matching_key(ground_truth, gt_sap, gt_tc)
    folder_matches = add_matching_key(folder_matches, match_sap, match_tc)
    folder_matches["_file_count"] = pd.to_numeric(
        folder_matches[match_file_count], errors="coerce"
    ).fillna(-1) if match_file_count is not None else -1

    # Multiple copies can exist under different source roots.  Pick the one
    # containing most files as the path used by the classifier.
    best_paths = (
        folder_matches.loc[folder_matches["_match_key"].ne("")]
        .sort_values(["_match_key", "_file_count"], ascending=[True, False])
        .drop_duplicates("_match_key", keep="first")
        .set_index("_match_key")[match_path]
    )

    def gt_value(column: object | None) -> pd.Series:
        if column is None:
            return pd.Series("", index=ground_truth.index)
        return ground_truth[column].map(clean_text)

    output = pd.DataFrame(
        {
            "SAP-Nummer": gt_value(gt_sap),
            "Teamcenter": gt_value(gt_tc),
            "Benennung (E)": gt_value(gt_english),
            "Benennung (D)": gt_value(gt_german),
            "Ground Truth": gt_value(gt_label),
            "Register": gt_value(gt_register),
            "Weitere zulässige Ground Truth": gt_value(gt_alternatives),
            "Data_Folder_Path": ground_truth["_match_key"].map(best_paths).fillna("").map(clean_text),
        }
    )

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(args.output, engine="openpyxl") as writer:
        output.to_excel(writer, sheet_name="Experiment_Dataset", index=False)

    fills = source_register_fills(args.ground_truth_excel, args.ground_truth_sheet, gt_register)
    workbook = load_workbook(args.output)
    worksheet = workbook["Experiment_Dataset"]
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    register_col = list(output.columns).index("Register") + 1
    for row in range(2, worksheet.max_row + 1):
        register_cell = worksheet.cell(row=row, column=register_col)
        register_cell.fill = copy(fills.get(clean_text(register_cell.value), PatternFill()))
        register_cell.font = Font(bold=True, color="7030A0" if register_cell.value == "M" else "000000")
        register_cell.alignment = Alignment(horizontal="center")

    widths = {
        "SAP-Nummer": 18,
        "Teamcenter": 18,
        "Benennung (E)": 24,
        "Benennung (D)": 24,
        "Ground Truth": 30,
        "Register": 12,
        "Weitere zulässige Ground Truth": 38,
        "Data_Folder_Path": 70,
    }
    for column_index, header in enumerate(output.columns, start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = widths[header]

    workbook.save(args.output)
    print(f"Created {args.output.resolve()} with {len(output)} Ground-Truth rows.")


if __name__ == "__main__":
    main()
