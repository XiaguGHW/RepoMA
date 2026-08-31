"""
Check whether the Baugruppen listed in an Excel file already have a matching
folder in one or more local or synced Bosch cloud locations.

The script only reads folders and writes Excel reports. It never copies, moves,
renames, or changes existing Baugruppen data.

Before the first run, paste your Windows folder paths into SEARCH_ROOTS below.
A Bosch OneDrive/Teams location must be available as a real local/synced path
in Windows Explorer; a browser URL cannot be scanned by Python.
"""

from __future__ import annotations

import argparse
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent

# ---------------------------------------------------------------------------
# Configuration: paste every folder that should be searched here.
# Each entry must be a Windows folder path, not a OneDrive/Teams web link.
# Example:
# r"C:\\Users\\NUL4FE\\OneDrive - Bosch Group\\...\\Datenset",
# ---------------------------------------------------------------------------
SEARCH_ROOTS: list[str] = [
    # r"C:\\path\\to\\your\\first\\Baugruppen_folder",
    # r"C:\\path\\to\\your\\synced_Bosch_OneDrive_folder",
]

DEFAULT_INPUT_EXCEL = SCRIPT_DIR / "129BG.xlsx"
DEFAULT_OUTPUT_DIR = SCRIPT_DIR

SAP_COLUMN_CANDIDATES = (
    "SAP-Nummer",
    "SAP Nummer",
    "SAP-Nummer ",
    "Baugruppennummer",
    "Baugruppen-ID",
    "Baugruppen_ID",
    "HBG",
)
TEAMCENTER_COLUMN_CANDIDATES = (
    "Teamcenter",
    "Teamcenter ID",
    "Teamcenter-ID",
    "Teamcenter Nummer",
    "Teamcenter-Nummer",
    "TC ID",
    "TC-ID",
)
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg"}
PDF_EXTENSIONS = {".pdf"}


def clean_identifier(value: object) -> str:
    """Return an Excel identifier as text, preserving IDs such as 0804FM8785."""
    if pd.isna(value):
        return ""
    text = str(value).strip()
    # Excel sometimes turns a purely numeric identifier into "123.0".
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def comparable_name(value: object) -> str:
    """Normalize names for exact folder/identifier comparison."""
    return re.sub(r"[\s_-]+", "", clean_identifier(value)).casefold()


def find_column(columns: Iterable[object], candidates: tuple[str, ...]) -> str | None:
    """Find a column name while tolerating spaces and upper/lower case."""
    normalized_columns = {
        clean_identifier(column).casefold(): str(column) for column in columns
    }
    for candidate in candidates:
        found = normalized_columns.get(candidate.casefold())
        if found:
            return found
    return None


def has_identifier_token(folder_name: str, identifier: str) -> bool:
    """Match an identifier embedded in a folder name, separated by non-alphanumerics."""
    if not identifier:
        return False
    pattern = rf"(?<![A-Z0-9]){re.escape(identifier.upper())}(?![A-Z0-9])"
    return re.search(pattern, folder_name.upper()) is not None


def validate_search_roots(raw_roots: Iterable[str]) -> tuple[list[Path], list[dict[str, str]]]:
    valid_roots: list[Path] = []
    invalid_roots: list[dict[str, str]] = []

    for raw_root in raw_roots:
        raw_root = str(raw_root).strip()
        if not raw_root:
            continue
        root = Path(raw_root).expanduser()
        if root.is_dir():
            valid_roots.append(root)
            print(f"[accessible] {root}")
        else:
            print(f"[not accessible] {raw_root}")
            invalid_roots.append(
                {
                    "configured_path": raw_root,
                    "status": "not_accessible",
                    "note": "Path does not exist or is not an accessible folder.",
                }
            )
    return valid_roots, invalid_roots


def index_folders(search_roots: Iterable[Path]) -> list[dict[str, str]]:
    """Recursively collect all folders once, so 129 BGs do not trigger 129 scans."""
    indexed: list[dict[str, str]] = []

    for root in search_roots:
        try:
            folders = [root]
            folders.extend(path for path in root.rglob("*") if path.is_dir())
        except OSError as error:
            print(f"Warning: could not fully scan {root}: {error}")
            continue

        for folder in folders:
            try:
                indexed.append(
                    {
                        "folder_name": folder.name,
                        "folder_path": str(folder.resolve()),
                        "source_root": str(root.resolve()),
                        "normalized_name": comparable_name(folder.name),
                    }
                )
            except OSError as error:
                print(f"Warning: could not read folder {folder}: {error}")

    return indexed


def find_matches(
    folder_index: list[dict[str, str]],
    sap_number: str,
    teamcenter_id: str,
) -> list[dict[str, str]]:
    """Return exact matches first; only use embedded matches if no exact match exists."""
    identifiers = (
        ("SAP-Nummer", sap_number),
        ("Teamcenter", teamcenter_id),
    )
    exact_matches: list[dict[str, str]] = []
    embedded_matches: list[dict[str, str]] = []

    for folder in folder_index:
        for identifier_type, identifier in identifiers:
            normalized_identifier = comparable_name(identifier)
            if normalized_identifier and folder["normalized_name"] == normalized_identifier:
                exact_matches.append(
                    {
                        **folder,
                        "match_type": f"exact_{identifier_type}",
                        "matched_identifier": identifier,
                    }
                )
                break

            if has_identifier_token(folder["folder_name"], identifier):
                embedded_matches.append(
                    {
                        **folder,
                        "match_type": f"embedded_{identifier_type}",
                        "matched_identifier": identifier,
                    }
                )
                break

    matches = exact_matches if exact_matches else embedded_matches

    # The same folder can match SAP and Teamcenter; retain one row only.
    unique_matches: dict[str, dict[str, str]] = {}
    for match in matches:
        unique_matches.setdefault(match["folder_path"], match)
    return list(unique_matches.values())


def folder_file_summary(folder_path: str) -> dict[str, int]:
    """Count basic file types. This is a structural check, not a content-quality check."""
    summary = {
        "file_count": 0,
        "image_count": 0,
        "pdf_count": 0,
    }
    try:
        for path in Path(folder_path).rglob("*"):
            if not path.is_file() or path.name.startswith("~$"):
                continue
            summary["file_count"] += 1
            suffix = path.suffix.casefold()
            if suffix in IMAGE_EXTENSIONS:
                summary["image_count"] += 1
            if suffix in PDF_EXTENSIONS:
                summary["pdf_count"] += 1
    except OSError as error:
        print(f"Warning: could not read files in {folder_path}: {error}")
    return summary


def build_report(
    dataframe: pd.DataFrame,
    sap_column: str,
    teamcenter_column: str,
    folder_index: list[dict[str, str]],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    report_rows: list[dict[str, object]] = []
    match_rows: list[dict[str, object]] = []

    for excel_row, (_, row) in enumerate(dataframe.iterrows(), start=2):
        sap_number = clean_identifier(row[sap_column])
        teamcenter_id = clean_identifier(row[teamcenter_column])
        matches = find_matches(folder_index, sap_number, teamcenter_id)

        report_row = row.to_dict()
        report_row["Excel_Row"] = excel_row
        report_row["SAP_used_for_search"] = sap_number
        report_row["Teamcenter_used_for_search"] = teamcenter_id

        if not sap_number and not teamcenter_id:
            report_row.update(
                {
                    "Folder_Status": "missing_identifier",
                    "Match_Count": 0,
                    "Match_Type": "",
                    "Found_By": "",
                    "Found_Folder_Paths": "",
                    "Source_Roots": "",
                    "Total_File_Count": 0,
                    "Total_Image_Count": 0,
                    "Total_PDF_Count": 0,
                }
            )
        elif not matches:
            report_row.update(
                {
                    "Folder_Status": "not_found",
                    "Match_Count": 0,
                    "Match_Type": "",
                    "Found_By": "",
                    "Found_Folder_Paths": "",
                    "Source_Roots": "",
                    "Total_File_Count": 0,
                    "Total_Image_Count": 0,
                    "Total_PDF_Count": 0,
                }
            )
        else:
            summaries = [folder_file_summary(match["folder_path"]) for match in matches]
            report_row.update(
                {
                    # A BG counts as found as soon as any matching folder exists.
                    # Several matching folders are kept only as information.
                    "Folder_Status": "found",
                    "Match_Count": len(matches),
                    "Match_Type": " | ".join(match["match_type"] for match in matches),
                    "Found_By": " | ".join(match["matched_identifier"] for match in matches),
                    "Found_Folder_Paths": "\n".join(match["folder_path"] for match in matches),
                    "Source_Roots": "\n".join(match["source_root"] for match in matches),
                    "Total_File_Count": sum(item["file_count"] for item in summaries),
                    "Total_Image_Count": sum(item["image_count"] for item in summaries),
                    "Total_PDF_Count": sum(item["pdf_count"] for item in summaries),
                }
            )
            for match, summary in zip(matches, summaries):
                match_rows.append(
                    {
                        "Excel_Row": excel_row,
                        "SAP-Nummer": sap_number,
                        "Teamcenter": teamcenter_id,
                        "folder_name": match["folder_name"],
                        "folder_path": match["folder_path"],
                        "source_root": match["source_root"],
                        "match_type": match["match_type"],
                        "matched_identifier": match["matched_identifier"],
                        **summary,
                    }
                )

        report_rows.append(report_row)

    return pd.DataFrame(report_rows), pd.DataFrame(match_rows)


def style_worksheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column_number, column_name in enumerate(worksheet.iter_cols(1, worksheet.max_column), start=1):
        header = worksheet.cell(row=1, column=column_number).value
        width = min(max(len(str(header or "")) + 2, 14), 28)
        if header in {"Found_Folder_Paths", "Source_Roots", "folder_path", "source_root"}:
            width = 65
        worksheet.column_dimensions[get_column_letter(column_number)].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_excel_report(
    report: pd.DataFrame,
    matches: pd.DataFrame,
    invalid_roots: list[dict[str, str]],
    output_path: Path,
) -> None:
    not_found = report[report["Folder_Status"].isin(["not_found", "missing_identifier"])].copy()
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        report.to_excel(writer, sheet_name="all_BG_check", index=False)
        not_found.to_excel(writer, sheet_name="not_found", index=False)
        matches.to_excel(writer, sheet_name="folder_matches", index=False)
        pd.DataFrame(invalid_roots).to_excel(
            writer, sheet_name="invalid_search_paths", index=False
        )

        for worksheet in writer.sheets.values():
            style_worksheet(worksheet)


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Check whether each BG in an Excel sheet has a matching folder."
    )
    parser.add_argument(
        "--input-excel",
        type=Path,
        default=DEFAULT_INPUT_EXCEL,
        help=f"Excel file; first sheet is used (default: same folder as script: {DEFAULT_INPUT_EXCEL.name})",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="Report folder (default: same folder as the script)",
    )
    parser.add_argument(
        "--search-root",
        action="append",
        default=None,
        help="Optional additional folder to search. Repeat this option for multiple paths.",
    )
    parser.add_argument(
        "--sap-column",
        default=None,
        help="Excel column containing the SAP number. Auto-detected by default.",
    )
    parser.add_argument(
        "--teamcenter-column",
        default=None,
        help="Excel column containing the Teamcenter ID. Auto-detected by default.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    input_excel = args.input_excel.expanduser()
    if not input_excel.is_file():
        raise FileNotFoundError(f"Input Excel file not found: {input_excel}")

    # A repeated --search-root can be used temporarily without editing SEARCH_ROOTS.
    raw_roots = [*SEARCH_ROOTS, *(args.search_root or [])]
    print(f"Configured search roots: {len(raw_roots)}")
    valid_roots, invalid_roots = validate_search_roots(raw_roots)
    if not valid_roots:
        raise ValueError(
            "No accessible search folders. Paste paths into SEARCH_ROOTS at the "
            "top of this script, or pass one or more --search-root arguments."
        )

    dataframe = pd.read_excel(input_excel, sheet_name=0, dtype=str)
    sap_column = args.sap_column or find_column(dataframe.columns, SAP_COLUMN_CANDIDATES)
    teamcenter_column = args.teamcenter_column or find_column(
        dataframe.columns, TEAMCENTER_COLUMN_CANDIDATES
    )
    if not sap_column or not teamcenter_column:
        raise ValueError(
            "Could not identify both SAP and Teamcenter columns. "
            f"Available columns: {list(dataframe.columns)}. "
            "Use --sap-column and --teamcenter-column if necessary."
        )

    print(f"Input Excel: {input_excel}")
    print(f"SAP column: {sap_column}")
    print(f"Teamcenter column: {teamcenter_column}")
    print(f"Accessible search roots: {len(valid_roots)}")
    print("Indexing folders once. This can take a while for large cloud-synced locations...")
    folder_index = index_folders(valid_roots)
    print(f"Indexed folders: {len(folder_index)}")

    report, matches = build_report(dataframe, sap_column, teamcenter_column, folder_index)

    output_dir = args.output_dir.expanduser()
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_path = output_dir / f"129BG_folder_coverage_{timestamp}.xlsx"
    write_excel_report(report, matches, invalid_roots, output_path)

    found_count = int((report["Folder_Status"] == "found").sum())
    not_found_count = int((report["Folder_Status"] == "not_found").sum())
    missing_identifier_count = int((report["Folder_Status"] == "missing_identifier").sum())
    print(f"BGs with at least one matching folder: {found_count}")
    print(f"BGs without a matching folder: {not_found_count}")
    if missing_identifier_count:
        print(f"BGs without SAP and Teamcenter identifiers: {missing_identifier_count}")
    print(f"Report written: {output_path}")


if __name__ == "__main__":
    main()
